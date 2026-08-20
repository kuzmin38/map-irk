"""Сводная таблица показаний в Excel — для инженера.

Он передаёт показания в сбытовую компанию, ему нужен файл, а не экран
в мессенджере. В таблице должны быть и счётчики без показаний: иначе
не видно, что осталось снять.
"""
import io

import pytest
from openpyxl import load_workbook

from bot import db, houses, report
from bot import handlers as H

PERIOD = '2026-08'


@pytest.fixture(autouse=True)
def baza(tmp_path, monkeypatch):
    monkeypatch.setattr(db, 'DB_PATH', str(tmp_path / 'test.db'))
    db.init()


@pytest.fixture
def dannye():
    d = next(h for h in houses.ALL_HOUSES if h['address'] == 'Седова 71')
    db.set_house_complex(d['id'], '4sun')
    dom = db.add_meter(d['id'], 'hvs', 'Подвал, ввод ХВС на дом', 'Андрей')
    db.update_meter(dom, serial='04517')
    db.add_reading(dom, 1000, '2026-07', 1, 'Андрей Кузьмин')
    db.add_reading(dom, 1234, PERIOD, 1, 'Андрей Кузьмин')
    ofis = db.add_meter(d['id'], 'hvs_office', 'Подвал, ввод на офисы', 'Андрей')
    return {'dom': dom, 'ofis': ofis, 'house': d}


def listy(data):
    return load_workbook(io.BytesIO(data)).active


def test_fayl_sobiraetsya_i_otkryvaetsya(dannye):
    ws = listy(report.meters_workbook(PERIOD, 'август 2026'))

    assert ws.title == 'Показания'
    assert 'август 2026' in str(ws['A1'].value)


def test_v_tablice_est_i_sdannye_i_nesdannye(dannye):
    rows = report.meters_rows(PERIOD)

    sdano = [r for r in rows if not r['_нет']]
    net = [r for r in rows if r['_нет']]
    assert len(sdano) == 1 and len(net) == 1
    assert 'офисы' in net[0]['Счётчик']


def test_raschyot_rashoda(dannye):
    r = next(r for r in report.meters_rows(PERIOD) if r['Счётчик'].endswith('на дом'))

    assert r['Прошлое'] == 1000
    assert r['Текущее'] == 1234
    assert r['Расход'] == 234


def test_zavodskoy_nomer_i_zhk_popadayut_v_fayl(dannye):
    r = next(r for r in report.meters_rows(PERIOD) if r['Заводской №'])

    assert r['Заводской №'] == '04517'
    assert 'солнца' in r['ЖК'].lower()
    assert r['Адрес'] == 'Седова 71'


def test_kto_podal_i_kogda(dannye):
    r = next(r for r in report.meters_rows(PERIOD) if not r['_нет'])

    assert r['Кто подал'] == 'Андрей Кузьмин'
    assert r['Когда']


def test_shapka_i_schyot_sdannyh(dannye):
    ws = listy(report.meters_workbook(PERIOD))

    assert 'Сдано 1 из 2' in str(ws['A2'].value)
    zagolovki = [c.value for c in ws[4]]
    assert zagolovki[:4] == ['ЖК', 'Адрес', 'Счётчик', 'Вид']


def test_nesdannaya_stroka_podsvechena(dannye):
    ws = listy(report.meters_workbook(PERIOD))

    cveta = {}
    for row in ws.iter_rows(min_row=5):
        cveta[row[2].value] = row[0].fill.fgColor.rgb
    assert cveta['Подвал, ввод на офисы'] != cveta['Подвал, ввод ХВС на дом']


def test_pustaya_baza_ne_lomaet_vygruzku():
    data = report.meters_workbook(PERIOD)

    ws = listy(data)
    assert 'Сдано 0 из 0' in str(ws['A2'].value)


def test_knopka_vygruzki_est_v_svodke():
    """Кнопка нужна инженеру — он передаёт показания в сбытовую."""
    assert 'engineer' in H.BRIEFING_ROLES
