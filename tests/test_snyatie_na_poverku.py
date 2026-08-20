"""Счётчик сняли на поверку — это состояние, а не догадка.

Заказчик: «сказали, что счётчик поставлен, а он не поставлен, искали
в столярке; приехал инспектор пломбировать, а прибора нет». Каждый шаг
теперь подписан: кто снял и когда, кто поставил и когда.
"""
import io

import pytest
from openpyxl import load_workbook

from bot import checks, db, houses, report
from bot import handlers as H


@pytest.fixture(autouse=True)
def baza(tmp_path, monkeypatch):
    monkeypatch.setattr(db, 'DB_PATH', str(tmp_path / 'test.db'))
    db.init()


@pytest.fixture
def schyotchik():
    d = houses.HOUSES[0]
    m = db.add_meter(d['id'], 'hvs', 'ХВС подвал', 'Андрей')
    db.update_meter(m, serial='04517')
    return m


def test_novyy_schyotchik_na_meste(schyotchik):
    assert db.get_meter(schyotchik)['status'] == db.METER_ACTIVE


def test_snyatie_zapisyvaet_kto_i_kogda(schyotchik):
    db.meter_remove(schyotchik, 5, 'Константин Толщин')

    m = db.get_meter(schyotchik)
    assert m['status'] == db.METER_REMOVED
    assert m['status_by'] == 'Константин Толщин'
    assert m['status_at']


def test_ustanovka_podpisyvaetsya_drugim_chelovekom(schyotchik):
    db.meter_remove(schyotchik, 5, 'Константин Толщин')
    db.meter_install(schyotchik, 7, 'Александр Палевич')

    m = db.get_meter(schyotchik)
    assert m['status'] == db.METER_ACTIVE
    assert m['status_by'] == 'Александр Палевич'


def test_zhurnal_hranit_vsyu_cepochku(schyotchik):
    db.meter_remove(schyotchik, 5, 'Константин Толщин', 'увезли в поверку')
    db.meter_install(schyotchik, 7, 'Александр Палевич')

    sobytiya = db.meter_events(schyotchik)

    assert [e['action'] for e in sobytiya] == [db.METER_ACTIVE, db.METER_REMOVED]
    assert sobytiya[1]['note'] == 'увезли в поверку'


def test_snyatye_sobirayutsya_spiskom(schyotchik):
    db.meter_remove(schyotchik, 5, 'Константин')

    assert [m['id'] for m in db.removed_meters()] == [schyotchik]

    db.meter_install(schyotchik, 7, 'Александр')
    assert db.removed_meters() == []


def test_snyatyy_podsvechen(schyotchik):
    db.meter_remove(schyotchik, 5, 'Константин Толщин')

    found = checks.house_findings(houses.HOUSES[0]['id'], '2026-08')

    pro_snyatyy = [f for f in found if 'снят на поверку' in f['text']]
    assert pro_snyatyy
    assert 'Константин Толщин' in pro_snyatyy[0]['text']


def test_dolgo_ne_vozvraschayut_krasnym(schyotchik, monkeypatch):
    """Месяц в поверке — уже не мелочь: инспектор приедет, а прибора нет."""
    db.meter_remove(schyotchik, 5, 'Константин')
    db.update_meter(schyotchik, status_at='01.01.2020 10:00')

    found = checks.house_findings(houses.HOUSES[0]['id'], '2026-08')

    assert any(f['level'] == checks.RED and 'снят на поверку' in f['text']
               for f in found)


def test_v_vygruzku_popadaet_sostoyanie(schyotchik):
    db.meter_remove(schyotchik, 5, 'Константин Толщин')

    rows = report.meters_rows('2026-08')

    assert 'СНЯТ НА ПОВЕРКУ' in rows[0]['Состояние']
    assert 'Константин Толщин' in rows[0]['Состояние']


def test_shapka_vygruzki_schitaet_snyatye(schyotchik):
    db.meter_remove(schyotchik, 5, 'Константин')

    ws = load_workbook(io.BytesIO(report.meters_workbook('2026-08'))).active

    assert 'снято на поверку: 1' in str(ws['A2'].value)


class FakeBody:
    def __init__(self, text):
        self.text, self.attachments, self.mid = text, [], 'm1'


class FakeRecipient:
    chat_type = 'dialog'
    chat_id = 1


class FakeSender:
    user_id = 5
    full_name = 'Андрей'


class FakeMessage:
    def __init__(self, text):
        self.body, self.recipient, self.sender = FakeBody(text), FakeRecipient(), FakeSender()
        self.sent = []

    async def answer(self, text=None, attachments=None):
        self.sent.append(text)


class FakeEvent:
    def __init__(self, text):
        self.message, self.bot = FakeMessage(text), None


async def test_pokazanie_po_snyatomu_ne_pishetsya(schyotchik, monkeypatch, tmp_path):
    """Прибора нет на месте — значит, и цифры с него взяться не могло."""
    monkeypatch.setattr(H, 'DOCS_DIR', str(tmp_path / 'docs'))
    db.meter_remove(schyotchik, 5, 'Константин Толщин')
    adres = houses.HOUSES[0]['address']
    e = FakeEvent(f'{adres} хвс 1234')

    assert await H.handle_readings(e, e.message.body.text, 5)

    assert db.meter_readings(schyotchik) == []
    assert 'снятым на поверку' in ' '.join(e.message.sent)
