"""Выгрузка показаний счётчиков в Excel.

Инженер передаёт показания в сбытовую компанию — ему нужен файл, а не
экран в мессенджере. В таблице есть и счётчики без показаний за месяц:
без них не видно, что осталось снять.
"""
import io
import logging

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from . import db, houses

log = logging.getLogger('report')

VIDY = {
    'hvs': 'ХВС — дом',
    'hvs_office': 'ХВС — офисы',
    'heat': 'Теплосчётчик',
    'gvs': 'ГВС',
    'other': 'Другой',
}

COLUMNS = [
    ('ЖК', 22), ('Адрес', 24), ('Счётчик', 34), ('Вид', 14),
    ('Заводской №', 16), ('Прошлое', 12), ('Текущее', 12), ('Расход', 11),
    ('Кто подал', 20), ('Когда', 18),
]

_HEAD = PatternFill('solid', fgColor='DCE6F1')
_MISSING = PatternFill('solid', fgColor='FCE4D6')


def meters_rows(period: str) -> list:
    """Строки выгрузки: по одной на каждый счётчик, включая не сданные."""
    complexes = db.all_house_complexes()
    names = {c['id']: c['name'] for c in houses.COMPLEXES}
    za_period = {r['meter_id']: r for r in db.readings_for_period(period)}
    rows = []
    for h in houses.HOUSES:
        for m in db.list_meters(h['id']):
            istoriya = db.meter_readings(m['id'], limit=12)
            tekushchee = za_period.get(m['id'])
            proshloe = next((r for r in istoriya if r['period'] != period), None)
            rashod = None
            if tekushchee and proshloe:
                rashod = round(tekushchee['value'] - proshloe['value'], 3)
            rows.append({
                'ЖК': names.get(complexes.get(h['id']), ''),
                'Адрес': h['address'],
                'Счётчик': m['label'],
                'Вид': VIDY.get(m['kind'], m['kind']),
                'Заводской №': m['serial'] or '',
                'Прошлое': proshloe['value'] if proshloe else None,
                'Текущее': tekushchee['value'] if tekushchee else None,
                'Расход': rashod,
                'Кто подал': (tekushchee['submitted_by_name'] or '') if tekushchee else '',
                'Когда': (tekushchee['submitted_at'] or '') if tekushchee else '',
                '_нет': tekushchee is None,
            })
    return rows


def meters_workbook(period: str, period_label: str = '') -> bytes:
    """Готовый файл .xlsx с показаниями за период."""
    rows = meters_rows(period)
    wb = Workbook()
    ws = wb.active
    ws.title = 'Показания'

    zagolovok = f'Показания счётчиков — {period_label or period}'
    ws.append([zagolovok])
    ws['A1'].font = Font(bold=True, size=13)
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(COLUMNS))
    sdano = sum(1 for r in rows if not r['_нет'])
    ws.append([f'Сдано {sdano} из {len(rows)} счётчиков'])
    ws.append([])

    head_row = ws.max_row + 1
    ws.append([name for name, _ in COLUMNS])
    for i in range(1, len(COLUMNS) + 1):
        cell = ws.cell(row=head_row, column=i)
        cell.font = Font(bold=True)
        cell.fill = _HEAD
        cell.alignment = Alignment(wrap_text=True, vertical='center')

    for r in rows:
        ws.append([r[name] for name, _ in COLUMNS])
        if r['_нет']:
            # Не сданные подсвечиваем: инженеру важно видеть пробелы
            for i in range(1, len(COLUMNS) + 1):
                ws.cell(row=ws.max_row, column=i).fill = _MISSING

    for i, (_, width) in enumerate(COLUMNS, start=1):
        ws.column_dimensions[get_column_letter(i)].width = width
    ws.freeze_panes = ws.cell(row=head_row + 1, column=1)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
